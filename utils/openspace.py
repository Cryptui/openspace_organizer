import random 
from typing import List 
from .table import Table 
import openpyxl

class Openspace:

    """
    A class to represent an open space.

    """
    def __init__(self, number_of_tables: int, table_capacity: int):
        """
        Initializes an Openspace object with the given number of tables and table capacity.
        Args:
            number_of_tables (int): The number of tables in the openspace.
            table_capacity (int): The maximum number of seats per table.
        """
 
        self.tables: List[Table] = [Table(table_capacity) for _ in range(number_of_tables)]
        self.number_of_tables: int = number_of_tables

    def organize(self, names: List[str]):
        """
        Randomly assigns people to seats in the openspace.
        Args:
            names (List[str]): A list of names of people to be assigned to seats.
        """
        # Shuffle the list of names randomly
        random.shuffle(names)

        # Flatten the list of seats
        all_seats = [seat for table in self.tables for seat in table.seat]

        # Check if the number of colleagues is less than the total capacity
        if len(names) <= len(all_seats):
            for name, seat in zip(names, all_seats):
                seat.set_occupant(name)
            return

        # Otherwise, distribute colleagues to seats as usual
        table_index = 0
        for name in names:
            # Find the next table with a free spot
            while table_index < self.number_of_tables and not self.tables[table_index].has_free_spot():
                table_index += 1
            # If all tables are full
            if table_index == self.number_of_tables:
                print("Not enough space to seat all colleagues.")
                return

            # Assign the seat to the person
            self.tables[table_index].assign_seat(name)

            # Move to the next table (circular)
            table_index = (table_index + 1) % self.number_of_tables

    def display(self):
        """Displays the seating arrangement in a readable format."""

        for i, table in enumerate(self.tables):
            print(f"Table {i+1}: ")
            for j, seat in enumerate(table.seat):
                print(f"\tSeat {j+1}: {'Empty' if seat.free else seat.occupant}")


    def store_seat_arrangement(self, filename: str):
        """
        Stores the seating arrangement in an Excel file.
        """
        try:
            # Add ".xlsx" extension if missing
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"

            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Seating Arrangement"

            for i, table in enumerate(self.tables):
                for j, seat in enumerate(table.seat):
                    sheet.cell(row=i * len(table.seat) + j + 1, column=1, value=f"Table {i + 1}")
                    sheet.cell(row=i * len(table.seat) + j + 1, column=2, value=f"Seat {j + 1}")
                    sheet.cell(row=i * len(table.seat) + j + 1, column=3, value=seat.occupant if not seat.free else "Empty")

            wb.save(filename)
            print(f"Seating arrangement saved to {filename} successfully.")
        except Exception as e:
            print(f"Error occurred while saving the seating arrangement: {e}")